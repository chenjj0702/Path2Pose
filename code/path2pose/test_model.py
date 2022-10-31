"""
基于 load_model 中的模型生成 所有 test_loader 中的数据
保存npz格式
绘制args.
"""
import torch
import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from tools.visualize import visualize_mul_seq


##
def cal_cc(_data1, _data2):
    out = np.zeros(len(_data1))
    for _i in range(len(_data1)):
        _a = pd.Series(_data1[_i])
        _b = pd.Series(_data2[_i])
        out[_i] = _b.corr(_a)
    return out


def get_traj_v(X, point):
    if X.shape[-1] == 44:
        X = X.reshape(X.shape[0], 22, 2)

    x = X[:, point, :]
    d_x = np.diff(x, axis=0)
    v = np.sqrt(np.sum(np.square(d_x), axis=1))  # (T-1,)

    x0 = X[0, point, :]
    s_x = x - x0  # ()
    s = np.sqrt(np.sum(np.square(s_x), axis=1))  # (T)
    return v, s


def get_S(data, point=0):
    if len(data.shape) == 3:
        data = data.reshape(data.shape[0], data.shape[1], 22, 2)
    V_list = []
    S_list = []
    for i, data_i in enumerate(data):
        V, S = get_traj_v(data_i, point)
        V_list.append(V)
        S_list.append(S)
    V_list = np.array(V_list)
    S_list = np.array(S_list)
    return S_list, V_list


def test(bone, loader, data, plot_num=0, form='.jpg'):
    gen = bone.gen
    args = bone.args
    device = bone.device
    plot_num = plot_num
    obj_dir = Path(args.obj_dir)
    # map_all = data['maps']
    gen.eval()

    out_list = []
    real_list = []
    cond_list = []
    map_list = []
    for batch_i, batch in enumerate(loader):
        # prepare data
        pose_glb, path_glb = map(lambda x: x.to(device), batch)

        guide_path = path_glb[:, 5:]
        gold_pose = pose_glb[:, 5:]
        init_pose = pose_glb[:, :5]

        with torch.no_grad():
            output = gen(guide_path, init_pose)

        output = torch.cat((init_pose, output), dim=1)
        gold_pose = torch.cat((init_pose, gold_pose), dim=1)
        out_list.append(output.detach().cpu().numpy())
        real_list.append(gold_pose.detach().cpu().numpy())
        cond_list.append(guide_path.detach().cpu().numpy())
        # map_list.extend(maps)

    if not obj_dir.exists():
        obj_dir.mkdir()

    out = np.concatenate(out_list, axis=0)  # (n,len_seq,d_pos)
    real = np.concatenate(real_list, axis=0)
    path = np.concatenate(cond_list, axis=0)  # (n,len_seq,2)
    if plot_num > 0:
        ids = np.arange(0, len(out), int(len(out) / plot_num))[:plot_num]
    else:
        ids = []

    # save
    npz_name = str(obj_dir) + '/test_results.npz'
    np.savez(npz_name, real=real, fake=out, path=path, plot_id=ids)

    ids = [104, 150]  # todo

    if plot_num > 0:
        real_plot = real[ids]
        path_plot = path[ids]
        real_plot = real_plot.reshape(
            (real_plot.shape[0], real_plot.shape[1], args.n_points, args.dim))
        visualize_mul_seq(real_plot, str(obj_dir), path_plot, ids, track_points=args.base_points, append='real', form=form)

        # fake
        fake_plot = out[ids]
        path_plot = path[ids]
        fake_plot = fake_plot.reshape(
            (fake_plot.shape[0], fake_plot.shape[1], args.n_points, args.dim))
        visualize_mul_seq(fake_plot, str(obj_dir), path_plot, ids, track_points=args.base_points, append='fake', form=form)

    # """ tail analysis """
    # tmp_dir = Path('C:/Users/cjj/Desktop/head_tail/')
    # if not tmp_dir.exists():
    #     tmp_dir.mkdir()
    #
    # S_real, _ = get_S(real)
    # S_fake, _ = get_S(out)
    # S_cc = cal_cc(S_real, S_fake)
    #
    # name = str(tmp_dir / 'tail_path_cc.xlsx')
    # wb = Workbook()
    # ws = wb.active
    # ws.append(['sample', 'S_cc'])
    # for i in range(len(S_cc)):
    #     ws.append([i + 1, S_cc[i]])
    # wb.save(name)
    #
    # name = str(tmp_dir / 'tail_example.xlsx')
    # wb = Workbook()
    # ws = wb.active
    # id_select = 1508
    # ws.append(['real', 'synthesized'])
    # s_real = S_real[id_select]
    # s_fake = S_fake[id_select]
    # for i in range(len(s_real)):
    #     ws.append([s_real[i], s_fake[i]])
    # ws.append([S_cc[id_select], 0])
    # wb.save(name)
    #
    # plt.rcParams['figure.figsize'] = [4, 3]
    # plt.rcParams['figure.dpi'] = 300
    # for i, id_i in enumerate(ids):
    #     # plot S
    #     ax = plt.subplot(1, 1, 1)
    #     plt.plot(S_real[id_i], 'o-', color='orange', markersize=3, label='real')
    #     plt.plot(S_fake[id_i], 'o-', color='blue', markersize=3, label='synthesized')
    #     plt.yticks([])
    #     plt.legend()
    #     ax.title.set_text(f'S cc={S_cc[id_i]:.3f}')
    #
    #     name = str(tmp_dir / ('tail_' + str(id_i) + '.jpg'))
    #     plt.savefig(name)
    #     plt.close()

    # 绘制临时路径
    # path_fake = fake_plot[:, :, 11, :]
    # path_cond = fake_plot[:, :5, 11, :]
    # path_real = np.concatenate((path_cond, path_plot), axis=1)
    #
    # plt.rcParams['figure.figsize'] = [3, 3]
    # plt.rcParams['figure.dpi'] = 300
    #
    # for i, id_i in enumerate(ids):
    #     real_i = path_real[i]
    #     fake_i = path_fake[i]
    #     plt.figure()
    #     plt.plot(real_i[:, 0], real_i[:, 1], '-', color='orange', label='real')
    #     plt.plot(fake_i[:, 0], fake_i[:, 1], 'o', color='blue', markersize=3, label='fake')
    #     plt.plot(fake_i[:5, 0], fake_i[:5, 1], 'o', color='green', markersize=3, label='conditional')
    #     plt.xticks([])
    #     plt.yticks([])
    #     plt.legend()
    #     name = str(tmp_dir / (str(id_i)+'.jpg'))
    #     plt.savefig(name)
