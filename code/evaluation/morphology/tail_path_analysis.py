import numpy as np
import matplotlib.pyplot as plt
from dtw import dtw
import scipy.interpolate as spi
from tqdm import tqdm
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from scipy.interpolate import UnivariateSpline
from scipy.signal import savgol_filter


##
def get_path(X, point):
    if X.shape[-1] == 44:
        X = X.reshape(X.shape[0], X.shape[1], 22, 2)  # (n,T,22,2)

    out = X[:, :, point, :]
    return out


def get_traj_v(X, point):
    if X.shape[-1] == 44:
        X = X.reshape(X.shape[0], 22, 2)

    x = X[:, point, :]
    d_x = np.diff(x, axis=0)
    v = np.sqrt(np.sum(np.square(d_x), axis=1))  # (T-1,)

    x0 = X[0, point, :]
    s_x = x - x0  # ()
    s = np.sqrt(np.sum(np.square(s_x), axis=1))  # (T)
    return v, s


def cal_dtw(a, b):
    # a = spline_interpolate(a)
    # b = spline_interpolate(b)

    # euclidean_norm = lambda x, y: np.abs(x - y)
    euclidean_norm = lambda x, y: np.sqrt(np.sum(np.square(x - y)))
    d, cost_matrix, acc_cost_matrix, path = dtw(a, b, dist=euclidean_norm)
    return d


def cal_cc(_data1, _data2):
    out = np.zeros(len(_data1))
    for _i in range(len(_data1)):
        _a = pd.Series(_data1[_i])
        _b = pd.Series(_data2[_i])
        out[_i] = _b.corr(_a)
    return out


def sm_fun(x, y, method):
    if method == 'spline':
        spl = UnivariateSpline(x, y, k=3)
        # spl.set_smoothing_factor(0.5)
        y_out = spl(x)
    elif method == 'SG':
        y_out = savgol_filter(y, 5, 2, mode='nearest')
    else:
        raise Exception
    return y_out


def sm_stack(_data):
    """
    :param _data: (n,T,dim)
    :return:
    """
    _data = np.swapaxes(_data, 1, 2)  # (n,dim,T)
    out = np.zeros_like(_data)  # (n,dim,T)
    for i, data_i in enumerate(_data):
        for j, line in enumerate(data_i):
            x = np.arange(len(line))
            tmp = sm_fun(x, line, 'SG')
            out[i, j, :] = tmp
    out = out.swapaxes(1, 2)  # (n,T,dim)
    return out


## main
if __name__ == '__main__':
    np.random.seed(1)
    npz_name = Path('../../../../results/path2pose/v11_public_attncn_220812_202737/val/epoch10000/test_results.npz')

    obj_dir = Path('../../../../results/path2pose/evaluation_morphology/tail_analysis')
    if not obj_dir.exists():
        obj_dir.mkdir(parents=True)

    load_data = np.load(npz_name)
    raw_real = load_data['real']
    raw_fake = load_data['fake']
    n, t, _ = raw_real.shape
    raw_real = raw_real.reshape(n, t, 22, 2)  # (n,t,22,2)
    raw_fake = raw_fake.reshape(n, t, 22, 2)

    # excel_name = '../../../../results/path2pose/assessment_morphology/raw_data_plot/直行.xlsx'
    # wb = load_workbook(excel_name)
    # ws = wb.active
    # select_ids = []
    # for row_i in ws.iter_rows():
    #     for cell in row_i:
    #         select_ids.append(cell.value)
    # select_ids = np.array(select_ids)

    select_ids = np.arange(len(raw_real))

    base_point = 0
    V_real, V_fake, S_real, S_fake = [], [], [], []
    for i, id_i in enumerate(select_ids):
        real_v, real_s = get_traj_v(raw_real[id_i], base_point)
        fake_v, fake_s = get_traj_v(raw_fake[id_i], base_point)

        V_real.append(real_v)
        V_fake.append(fake_v)
        S_real.append(real_s)
        S_fake.append(fake_s)
    S_real = np.array(S_real)
    S_fake = np.array(S_fake)
    V_real = np.array(V_real)[:, :, np.newaxis]  # (n,t,1)
    V_fake = np.array(V_fake)[:, :, np.newaxis]
    V_real = sm_stack(V_real).squeeze()
    V_fake = sm_stack(V_fake).squeeze()

    """ 相关系数 """
    S_cc = cal_cc(S_real, S_fake)  # (n,)
    V_cc = cal_cc(V_real, V_fake)

    name = str(obj_dir / 'tail_path_cc.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.append(['sample', 'S_cc', 'V_cc'])
    for i in range(len(S_cc)):
        ws.append([i + 1, S_cc[i], V_cc[i]])
    wb.save(name)

    """ plot sample """
    plt.rcParams['figure.figsize'] = [6, 2]
    plt.rcParams['figure.dpi'] = 300
    ids = np.random.choice(len(S_real), 50)

    for i, id_i in enumerate(ids):
        # plot S
        ax = plt.subplot(1, 2, 1)
        plt.plot(S_real[id_i], 'o-', color='orange', markersize=3, label='real')
        plt.plot(S_fake[id_i], 'o-', color='blue', markersize=3, label='fake')
        plt.yticks([])
        plt.legend()
        ax.title.set_text(f'S cc={S_cc[id_i]:.3f}')

        # plot V
        ax = plt.subplot(1, 2, 2)
        plt.plot(V_real[id_i], 'o-', color='orange', markersize=3, label='real')
        plt.plot(V_fake[id_i], 'o-', color='blue', markersize=3, label='fake')
        plt.yticks([])
        plt.legend()
        ax.title.set_text(f'V cc={V_cc[id_i]:.3f}')

        name = str(obj_dir / ('tail_' + str(id_i) + '.jpg'))
        plt.savefig(name)
        plt.close()
