import numpy as np
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.pyplot import MultipleLocator

matplotlib.use('Qt5Agg')
import sys

sys.path.append('../../utils')
from animation_larva import plot_pose_seq
from tqdm import tqdm
from scipy.stats import pearsonr


##
def cal_mid_line(_data):
    n, t, _ = _data.shape
    _data = _data.reshape(-1, _data.shape[-1])  # (n,44)

    pose = _data.reshape(len(_data), 22, 2)  # (n,22,2)
    tmp1 = pose[:, :12, :]  # (0-11)  (n,12,2)
    tmp2 = np.concatenate((pose[:, 11:, :], pose[:, [0], :]), axis=1)  # (11-21,0)  # (n,12,2)
    tmp2 = np.flip(tmp2, axis=1)  # (0,21-11)
    pose_new = np.stack((tmp1, tmp2), axis=1)  # (n,2,12,c)

    mid = pose_new.mean(axis=1)  # (n,12,c)

    # test
    # test_id = 221
    # a = mid[test_id]  # (12,2)
    # b = pose[test_id]
    # plt.figure()
    # x_min, x_max, y_min, y_max = b[:, 0].min(), b[:, 0].max(), b[:, 1].min(), b[:, 1].max()
    # r_x, r_y = x_max - x_min, y_max - y_min
    # r = max(r_x, r_y)
    # plt.axis([x_min, x_min+r, y_min, y_min+r])
    # plt.plot(a[:, 0], a[:, 1], '.-')
    # plt.plot(b[:, 0], b[:, 1], '.-')
    # plt.show()

    # 计算距离
    m_dif = np.diff(mid, axis=1)  # (n,11,c)
    m_dist = np.sqrt(np.sum(np.square(m_dif), axis=-1))  # (n,11)
    m_dist = m_dist.reshape(n, t, m_dist.shape[-1])

    # 计算角度
    n_angle = m_dif.shape[1] - 1
    v1 = mid[:, 1:, :] - mid[:, :-1, :]
    v1 = v1[:, 1:]  # (n,11,2)
    v2 = mid[:, :-1, :] - mid[:, 1:, :]
    v2 = v2[:, :-1]  # (n,11,2)

    v1 = v1.reshape(-1, 2)  # (n*11, 2)
    v2 = v2.reshape(-1, 2)  # (n*11, 2)

    angle1 = np.arctan2(v1[:, 1], v1[:, 0])  # (N,)
    angle2 = np.arctan2(v2[:, 1], v2[:, 0])  # (N,)

    tmp_id = np.argwhere(angle1 < angle2)
    angle1[tmp_id] = angle1[tmp_id] + np.pi * 2

    angle = angle1 - angle2  # (N)
    # angle = angle * 180 / np.pi
    angle = angle.reshape(len(m_dif), n_angle)
    # print(angle[test_id])
    angle = angle.reshape(n, t, angle.shape[-1])

    return m_dist, angle


def scale_dist(_data):
    _n, _t, _dim = _data.shape
    _data = _data.reshape(_n * _t, _data.shape[-1])  # (N,11)

    L = np.sum(_data, axis=-1)  # (N,)
    L = L[:, np.newaxis].repeat(_dim, axis=-1)
    out = _data / L
    out = out.reshape(_n, _t, _dim)
    return out


def plot_frame(x):
    # x (24,n_dim)
    assert x.ndim == 2 and x.shape[0] == 22 and x.shape[1] == 2

    # lines
    line_list = [[0, 1], [1, 2], [2, 3], [3, 4], [4, 5], [5, 6], [6, 7], [7, 8], [8, 9], [9, 10], [10, 11], [11, 12],
                 [12, 13], [13, 14], [14, 15], [15, 16], [16, 17], [17, 18], [18, 19], [19, 20], [20, 21], [21, 0],
                 [1, 21], [2, 20], [3, 19], [4, 18], [5, 17], [6, 16], [7, 15], [8, 14], [9, 13], [10, 12]]
    lines = []
    for pair in line_list:
        l_i, = plt.plot(x[pair, 0], x[pair, 1], 'cyan', alpha=0.8, linewidth=2)
        lines.append(l_i)

    # dots
    dots = []
    for i_, x_i in enumerate(x):
        if i_ == 0:
            dot, = plt.plot(x_i[0], x_i[1], color='red', marker='o', markersize=8)
        elif i_ == 11:
            dot, = plt.plot(x_i[0], x_i[1], color='green', marker='o', markersize=8)
        else:
            dot, = plt.plot(x_i[0], x_i[1], color='black', marker='o', markersize=5, alpha=1)
        dots.append(dot)
    return dots, lines, line_list


def plot_seq(_pose, _dist, _angle, _save_dir, base_point, r_disc, r_angle):
    minx, maxx = np.min(_pose[:, :, 0]), np.max(_pose[:, :, 0])
    miny, maxy = np.min(_pose[:, :, 1]), np.max(_pose[:, :, 1])
    deltx = maxx - minx
    delty = maxy - miny
    delt = max(deltx, delty)
    axis = [minx, minx + delt, miny, miny + delt]

    for _i, _pose_i in enumerate(_pose):
        _name = _save_dir / (str(_i) + '.jpg')
        plt.rcParams['figure.figsize'] = [10, 6]
        plt.rcParams['figure.dpi'] = 150
        # pose
        plt.subplot(2, 3, 1)
        plt.axis('off')
        plt.xticks([])
        plt.yticks([])
        plt.axis(axis)

        plt.plot(_pose[:, base_point, 0], _pose[:, base_point, 1], 'go', markersize=3)  # 整体路径
        plot_frame(_pose_i)

        # dist
        for _j, _disc_i in enumerate(_dist.T):
            ax = plt.subplot(5, 3, _j * 3 + 2)
            plt.xticks([])
            # plt.yticks([])
            # plt.ylim(r_disc[_j])
            plt.plot(range(len(_disc_i)), _disc_i, '-o', color='k', alpha=0.5)
            plt.plot(_i, _disc_i[_i], 'o', color='r')
            if _j == 0:
                ax.title.set_text('mid_line_distance')

        # angle
        for _j, _angle_i in enumerate(_angle.T):
            ax = plt.subplot(5, 3, _j * 3 + 3)
            plt.xticks([])
            # plt.yticks([])
            # plt.ylim(r_angle[_j])
            plt.plot(range(len(_angle_i)), _angle_i, '-o', color='k', alpha=0.5)
            plt.plot(_i, _angle_i[_i], 'o', color='r')
            if _j == 0:
                ax.title.set_text('mid_line_angle')
        plt.savefig(_name)
        plt.close()


def plot_seq_seperate(_pose, _dist, _angle, _save_dir, base_point):
    minx, maxx = np.min(_pose[:, :, 0]), np.max(_pose[:, :, 0])
    miny, maxy = np.min(_pose[:, :, 1]), np.max(_pose[:, :, 1])
    deltx = maxx - minx
    delty = maxy - miny
    delt = max(deltx, delty)
    axis = [minx, minx + delt, miny, miny + delt]

    for _i, _pose_i in enumerate(_pose):
        _name = _save_dir / (str(_i) + '.jpg')
        plt.rcParams['figure.figsize'] = [6, 6]
        plt.rcParams['figure.dpi'] = 300
        # pose
        plt.figure()
        plt.subplots_adjust(left=0, right=1, bottom=0, top=1)
        # plt.axis('off')
        plt.xticks([])
        plt.yticks([])
        plt.axis(axis)

        plt.plot(_pose[:, base_point, 0], _pose[:, base_point, 1], 'go', markersize=3)  # 整体路径
        plot_frame(_pose_i)
        plt.savefig(_name)
        plt.close()

    y_min, y_max = _dist.min(), _dist.max()
    for _j, _disc_i in enumerate(_dist.T):
        plt.figure(figsize=(5, 3))
        plt.ylim((y_min, y_max))
        plt.xticks([])
        # plt.yticks([])
        # plt.ylim(r_disc[_j])
        plt.plot(range(len(_disc_i)), _disc_i, '-o', color='b', alpha=0.8)

        _name = _save_dir / ('mid_line_pc' + str(_j) + '.jpg')
        plt.savefig(_name)
        plt.close()

    # angle
    y_min, y_max = _angle.min(), _angle.max()
    for _j, _angle_i in enumerate(_angle.T):
        plt.figure(figsize=(5, 3))
        plt.ylim((y_min, y_max))
        plt.xticks([])
        # plt.yticks([])
        # plt.ylim(r_angle[_j])
        plt.plot(range(len(_angle_i)), _angle_i, '-o', color='b', alpha=0.8)

        _name = _save_dir / ('angle_pc' + str(_j) + '.jpg')
        plt.savefig(_name)
        plt.close()


def plot_compare(_disc_pca, _angle_pca, _disc_cc, _angle_cc, _name):
    n_pc = len(_disc_pca[0].T)
    plt.rcParams['figure.figsize'] = [9, 6]
    plt.rcParams['figure.dpi'] = 300
    for _i in range(n_pc):
        # dist pca
        ax = plt.subplot(n_pc, 2, _i * 2 + 1)
        plt.subplots_adjust(wspace=0.3, hspace=0.5)
        plt.plot(_disc_pca[0][:, _i], 'o-', color='red', markersize=3)
        plt.plot(_disc_pca[1][:, _i], 'o-', color='blue', markersize=3)
        if _i < n_pc - 1:
            plt.xticks([])
        ax.title.set_text(f'distance PC{_i} cc={_disc_cc[_i]:.3f}')

        # angle pca
        plt.subplot(n_pc, 2, _i * 2 + 2)
        ax = plt.subplot(n_pc, 2, _i * 2 + 2)
        plt.plot(_angle_pca[0][:, _i], 'o-', color='red', markersize=3)
        plt.plot(_angle_pca[1][:, _i], 'o-', color='blue', markersize=3)
        if _i < n_pc - 1:
            plt.xticks([])
        ax.title.set_text(f'angle PC{_i} cc={_angle_cc[_i]:.3f}')

    plt.savefig(_name)
    plt.close()


## main
if __name__ == '__main__':
    obj_dir = Path('../../../../results/path2pose/evaluation_morphology/eigen')
    if not obj_dir.exists():
        obj_dir.mkdir(parents=True)

    np.random.seed(1)
    npz_name = Path(
        '../../../../results/path2pose/v5_GenAttnCn_DiscCnnSN_predict_5_35_nosm_for_assess_220328_154133/test/40000/test_results.npz')
    load_data = np.load(npz_name)
    raw_real = load_data['real']
    raw_fake = load_data['fake']
    mid_dist_real, mid_angle_real = cal_mid_line(raw_real)
    mid_dist_fake, mid_angle_fake = cal_mid_line(raw_fake)
    mid_dist_real = scale_dist(mid_dist_real)  # 归一化
    mid_dist_fake = scale_dist(mid_dist_fake)  # 归一化
    n, t, _ = mid_dist_real.shape

    """ eigenwave """
    y_disc_real = mid_dist_real.reshape(n * t, -1)  # (N,10)
    y_disc_fake = mid_dist_fake.reshape(n * t, -1)
    PCA_disc = PCA()
    PCA_disc.fit(y_disc_real)
    disc_pca_real = PCA_disc.transform(y_disc_real)
    disc_pca_fake = PCA_disc.transform(y_disc_fake)
    disc_pca_real = disc_pca_real.reshape(n, t, -1)
    disc_pca_fake = disc_pca_fake.reshape(n, t, -1)
    disc_pca_lim = [[x.min(), x.max()] for x in disc_pca_real.T]
    plt.rcParams['savefig.dpi'] = 300
    plt.rcParams['font.size'] = 12
    plt.figure()
    plt.plot(np.cumsum(PCA_disc.explained_variance_ratio_), 'o-')
    x_major_locator = MultipleLocator(1)
    ax = plt.gca()
    ax.xaxis.set_major_locator(x_major_locator)
    plt.xlabel('PC')
    plt.ylabel('variance')
    plt.title('eigenwave')
    # plt.savefig(str(obj_dir)+'eigenwave_variance.jpg')

    # excel_name = str(obj_dir / 'variance.xlsx')
    # wave_variance = np.cumsum(PCA_disc.explained_variance_ratio_)
    # wave_variance = pd.DataFrame(wave_variance)
    # wave_variance.to_excel(excel_name, sheet_name='eigenwave', index=False, header='variance')

    """ eigenbody """
    y_angle_real = mid_angle_real.reshape(n * t, -1)
    y_angle_fake = mid_angle_fake.reshape(n * t, -1)
    PCA_angle = PCA()
    PCA_angle.fit(y_angle_real)
    angle_pca_real = PCA_angle.transform(y_angle_real)
    angle_pca_fake = PCA_angle.transform(y_angle_fake)
    angle_pca_real = angle_pca_real.reshape(n, t, -1)
    angle_pca_fake = angle_pca_fake.reshape(n, t, -1)
    angle_pca_lim = [[x.min(), x.max()] for x in angle_pca_real.T]
    plt.rcParams['savefig.dpi'] = 300
    plt.rcParams['font.size'] = 12
    plt.figure()
    plt.plot(np.cumsum(PCA_angle.explained_variance_ratio_), 'o-')
    x_major_locator = MultipleLocator(1)
    ax = plt.gca()
    ax.xaxis.set_major_locator(x_major_locator)
    plt.xlabel('PC')
    plt.ylabel('variance')
    plt.title('eigenbody')
    # plt.savefig(str(obj_dir)+'eigenbody_variance.jpg')

    # body_variance = np.cumsum(PCA_angle.explained_variance_ratio_)
    # body_variance = pd.DataFrame(body_variance)
    # writer = pd.ExcelWriter(excel_name, engine='openpyxl', mode='a')
    # body_variance.to_excel(writer, sheet_name='eigenbody', index=False, header='variance')
    # writer.save()
    # writer.close()

    """ plot """
    n_plot = 200
    ids = np.random.choice(len(raw_real), n_plot, replace=False)
    samples_dir = obj_dir / 'samples'

    """ plot for paper """
    # raw_real = raw_real.reshape(n, t, 22, 2)
    # for i, id_i in enumerate(tqdm(ids)):
    #     # if id_i != 1548:
    #     #     continue
    #     dir_i = samples_dir / str(id_i).zfill(3)
    #     if not dir_i.exists():
    #         dir_i.mkdir(parents=True)
    #
    #     pose_i = raw_real[id_i]  # (t,22,2)
    #     disc_pca_i = disc_pca_real[id_i, :, :4]  # (t,5)
    #     angle_pca_i = angle_pca_real[id_i, :, :4]  # (t,5)
    #
    #     plot_seq_seperate(pose_i, disc_pca_i, angle_pca_i, dir_i, 11)
    #
    #     name = dir_i / 'eigen.xlsx'
    #     pd_wave = pd.DataFrame(disc_pca_i)
    #     pd_body = pd.DataFrame(angle_pca_i)
    #     pd_wave.to_excel(name, sheet_name='eigenwave', header=['PC1', 'PC2', 'PC3', 'PC4'], index=False)
    #     writer = pd.ExcelWriter(name, engine='openpyxl', mode='a')
    #     pd_body.to_excel(writer, sheet_name='eigenbody', header=['PC1', 'PC2', 'PC3', 'PC4'], index=False)
    #     writer.save()
    #     writer.close()

    """ all in one figure """
    # real
    # obj_dir = Path('../../../../results/path2pose/assessment_pca_mid/figure_real_40_3')
    # if not obj_dir.exists():
    #     obj_dir.mkdir(parents=True)
    # raw_real = raw_real.reshape(n, t, 22, 2)
    # for i, id_i in enumerate(tqdm(ids)):
    #     dir_i = obj_dir / str(id_i).zfill(3)
    #     if not dir_i.exists():
    #         dir_i.mkdir(parents=True)
    #
    #     pose_i = raw_real[id_i]  # (t,22,2)
    #     disc_pca_i = disc_pca_real[id_i, :, :5]  # (t,5)
    #     angle_pca_i = angle_pca_real[id_i, :, :5]  # (t,5)
    #
    #     plot_seq(pose_i, disc_pca_i, angle_pca_i, dir_i, 11, disc_pca_lim, angle_pca_lim)

    # fake
    # obj_dir = Path('../../../../results/path2pose/assessment_pca_mid/figure_fake')
    # if not obj_dir.exists():
    #     obj_dir.mkdir(parents=True)
    # raw_fake = raw_fake.reshape(n, t, 22, 2)
    # for i, id_i in enumerate(tqdm(ids)):
    #     dir_i = obj_dir / str(id_i).zfill(3)
    #     if not dir_i.exists():
    #         dir_i.mkdir(parents=True)
    #
    #     pose_i = raw_fake[id_i]  # (t,22,2)
    #     disc_pca_i = disc_pca_fake[id_i, :, :5]  # (t,5)
    #     angle_pca_i = angle_pca_fake[id_i, :, :5]  # (t,5)
    #
    #     plot_seq(pose_i, disc_pca_i, angle_pca_i, dir_i, 0, disc_pca_lim, angle_pca_lim)

    """ statistic """
    # real vs fake
    stat_dir = obj_dir / 'statistic'
    if not stat_dir.exists():
        stat_dir.mkdir(parents=True)
    disc_cc_pool, angle_cc_pool = [], []
    for i, id_i in enumerate(tqdm(ids)):
        disc_real_pca_i = disc_pca_real[id_i, :, :4]
        disc_fake_pca_i = disc_pca_fake[id_i, :, :4]

        disc_cc = np.zeros((len(disc_real_pca_i.T), 2))
        for j in range(len(disc_real_pca_i.T)):
            a = disc_real_pca_i[:, j]
            b = disc_fake_pca_i[:, j]
            disc_cc[j] = pearsonr(a, b)
        disc_cc_pool.append(disc_cc)

        angle_real_pca_i = angle_pca_real[id_i, :, :4]
        angle_fake_pca_i = angle_pca_fake[id_i, :, :4]

        angle_cc = np.zeros((len(angle_real_pca_i.T), 2))
        for j in range(len(angle_real_pca_i.T)):
            a = angle_real_pca_i[:, j]
            b = angle_fake_pca_i[:, j]
            angle_cc[j] = pearsonr(a, b)
        angle_cc_pool.append(angle_cc)

        # fig_name = stat_dir / (str(id_i) + '.jpg')
        # plot_compare([disc_real_pca_i, disc_fake_pca_i], [angle_real_pca_i, angle_fake_pca_i],
        #              disc_cc, angle_cc, fig_name)
        #
        # excel_name = stat_dir / (str(id_i) + '.xlsx')
        # header = ['real_1', 'real_2', 'real_3', 'real_4', 'fake_1', 'fake_2', 'fake_3', 'fake_4']
        # pd_dist = np.concatenate((disc_real_pca_i, disc_fake_pca_i), axis=1)
        # pd_dist = pd.DataFrame(pd_dist)
        # pd_angle = np.concatenate((angle_real_pca_i, angle_fake_pca_i), axis=1)
        # pd_angle = pd.DataFrame(pd_angle)
        #
        # pd_dist.to_excel(excel_name, sheet_name='eigenwave', header=header, index=True)
        # writer = pd.ExcelWriter(excel_name, engine='openpyxl', mode='a')
        # pd_angle.to_excel(writer, sheet_name='eigenbody', header=header, index=True)
        # writer.save()
        # writer.close()

    # save results for origin plot
    disc_cc_pool = np.array(disc_cc_pool)  # (n,4,2)
    angle_cc_pool = np.array(angle_cc_pool)  # (n,4,2)
    eigen_wave_cc = disc_cc_pool[:, :, 0].tolist()
    eigen_body_cc = angle_cc_pool[:, :, 0].tolist()
    eigen_wave_p = (disc_cc_pool[:, :, 1] < 0.05).sum(0)
    eigen_body_p = (angle_cc_pool[:, :, 1] < 0.05).sum(0)
    print(f'number of significant correlation for eigenwave1-4', eigen_wave_p)
    print(f'number of significant correlation for eigenbody1-4', eigen_body_p)

    excel_name = str(obj_dir / 'cc.xlsx')
    wb = Workbook()
    ws1 = wb.create_sheet('eigenwave_cc')
    ws2 = wb.create_sheet('eigenbody_cc')
    ws1.append(['PC1', 'PC2', 'PC3', 'PC4'])
    ws2.append(['PC1', 'PC2', 'PC3', 'PC4'])
    for row_i in eigen_wave_cc:
        ws1.append(row_i)
    for row_i in eigen_body_cc:
        ws2.append(row_i)
    wb.save(excel_name)
