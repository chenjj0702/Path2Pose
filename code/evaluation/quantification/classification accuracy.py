import numpy as np
import torch
import argparse
import csv
from tqdm import tqdm
from pathlib import Path
from torch.utils.data import Dataset
from classifier import Classifier_Rnn, Classifier_Cnn_40
import torch.nn.functional as F
import torch.optim as optim
from openpyxl import load_workbook, Workbook
from enhance import enh


## fun
class DanceDataset(Dataset):
    def __init__(self, dances, labels=None, detail_labels=None):
        if labels is not None:
            assert (len(labels) == len(dances)), \
                'the number of dances should be equal to the number of labels'
        self.dances = dances
        self.labels = labels
        self.detail_labels = detail_labels

    def __len__(self):
        return len(self.dances)

    def __getitem__(self, index):
        if self.labels is not None:
            return self.dances[index], self.labels[index], self.detail_labels[index]
        else:
            return self.dances[index]


def prepare_dataloader(dance_data, labels, detail_labels, batch):
    data_loader = torch.utils.data.DataLoader(
        DanceDataset(dance_data, labels, detail_labels),
        # num_workers=8,
        batch_size=batch,
        shuffle=True,
        # collate_fn=paired_collate_fn,
        pin_memory=True
    )

    return data_loader


def select_pose(x, train_rate):
    n_train = int(len(x) * train_rate)
    np.random.seed(1)
    train_index = np.random.choice(len(x), n_train, replace=False)
    test_index = np.array(list(set(np.arange(len(x))).difference(train_index)))

    train_x = x[train_index]
    test_x = x[test_index]
    return train_x, test_x


def split_seq(_x):
    _n, _t, _ = _x.shape
    if _t > 32:
        out = []
        start_i = 0
        end_i = 32
        step = 31
        while end_i < _t:
            out.append(_x[:, start_i:end_i, :])
            start_i += step
            end_i += step
        out = np.concatenate(out, axis=0)
    else:
        out = _x
    return out


def build_data(root, file_list, enhance=False, train_rate=0.8):
    train_pose_list, train_label_list, train_label2_list = [], [], []
    test_pose_list, test_label_list, test_label2_list = [], [], []

    for i, file_i in enumerate(file_list):
        file_i = str(root / file_i)
        load_i = np.load(file_i)

        if i == 0:
            real_tmp = load_i['real']
            # real_tmp = split_seq(real_tmp)  # todo 超过32帧自动剪裁
            train_pose, test_pose = select_pose(real_tmp, train_rate)
            if enhance:
                train_pose, test_pose = enh(train_pose, 8), enh(test_pose, 8)
            train_label = np.ones((len(train_pose, )))  # 真假label
            test_label = np.ones(len(test_pose))
            train_label2 = np.ones((len(train_pose, ))) * i  # 详细label 0 为真
            test_label2 = np.ones(len(test_pose)) * i

            train_pose_list.append(train_pose)
            test_pose_list.append(test_pose)
            train_label_list.append(train_label)
            test_label_list.append(test_label)
            train_label2_list.append(train_label2)
            test_label2_list.append(test_label2)

        fake_tmp = load_i['fake']
        # fake_tmp = split_seq(fake_tmp)  # todo 超过32帧自动剪裁
        train_pose, test_pose = select_pose(fake_tmp, train_rate)
        if enhance:
            train_pose, test_pose = enh(train_pose, 8), enh(test_pose, 8)
        train_label = np.zeros((len(train_pose, )))
        test_label = np.zeros(len(test_pose, ))
        train_label2 = np.ones(len(train_pose)) * (i + 1)
        test_label2 = np.ones(len(test_pose)) * (i + 1)

        train_pose_list.append(train_pose)
        test_pose_list.append(test_pose)
        train_label_list.append(train_label)
        test_label_list.append(test_label)
        train_label2_list.append(train_label2)
        test_label2_list.append(test_label2)

    train_pose = np.concatenate(train_pose_list, axis=0)
    train_label = np.concatenate(train_label_list, axis=0)
    test_pose = np.concatenate(test_pose_list, axis=0)
    test_label = np.concatenate(test_label_list, axis=0)
    train_label2 = np.concatenate(train_label2_list, axis=0)
    test_label2 = np.concatenate(test_label2_list, axis=0)
    save_test = np.stack(test_pose_list, axis=0)  # 保存test数据，用于计算FID

    return train_pose, train_label, train_label2, test_pose, test_label, test_label2, save_test


def val(dev_data, classifier, device, args):
    classifier.eval()
    corrects, avg_loss, size = 0, 0, 0
    for i, batch in enumerate(dev_data):
        dance, label, detail_label = map(lambda x: x.to(device), batch)
        dance = dance.float()
        label = label.long()

        pred, _ = classifier(dance)
        loss = F.cross_entropy(pred, label, args.w)
        avg_loss += loss.item()
        corrects += (torch.max(pred, 1)[1].view(label.size()).data == label.data).sum()
        size += batch[0].shape[0]

    avg_loss /= size
    accuracy = 100.0 * corrects / size
    print('Evaluation - loss: {:.6f}  acc: {:.4f}%({}/{}) \n'.format(avg_loss,
                                                                     accuracy,
                                                                     corrects,
                                                                     size))
    return accuracy


def train(train_data, dev_data, classifier, mode_i, _args):
    optimizer = optim.Adam(filter(lambda x: x.requires_grad, classifier.parameters()), lr=_args.lr)
    steps = 0
    val_train = []
    val_test = []
    for epoch in range(1, _args.epochs + 1):
        classifier.train()
        train_pool = []
        for i, batch in enumerate(train_data):
            dance, label, _ = map(lambda x: x.to(_args.device), batch)
            label = label.long()
            dance = dance.float()
            optimizer.zero_grad()
            pred, _ = classifier(dance)
            loss = F.cross_entropy(pred, label, weight=_args.w)
            loss.backward()
            optimizer.step()
            steps += 1

            corrects = (torch.max(pred, 1)[1].view(label.size()).data == label.data).sum()
            train_acc = 100.0 * corrects / batch[0].shape[0]
            train_pool.append(train_acc.detach().cpu().numpy())

        train_pool = np.array(train_pool).mean()
        val_train.append(train_pool)
        print(f'epoch{epoch} - loss: {loss:.6f} - acc: {train_pool:.6f}')
        _wb = load_workbook(args.log_train)
        _ws = _wb.active
        if mode_i == 0:
            _ws.cell(row=epoch + 1, column=1, value=epoch)
        _ws.cell(row=epoch + 1, column=mode_i + 2, value=train_pool)
        _wb.save(args.log_train)

        # evaluate the model on test set at each epoch
        dev_acc = val(dev_data, classifier, _args.device, _args)
        _wb = load_workbook(args.log_test)
        _ws = _wb.active
        if mode_i == 0:
            _ws.cell(row=epoch + 1, column=1, value=epoch)
        _ws.cell(row=epoch + 1, column=mode_i + 2, value=float(dev_acc.detach().cpu().numpy()))
        _wb.save(args.log_test)

        # if epoch % _args.save_interval == 0:
        #     name = _args.save_dir / ('epoch' + str(epoch) + '.pt')
        #     torch.save(classifier, name)


## main
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--batch_size', type=int, default=256)
    parser.add_argument('--epochs', type=int, default=100)
    parser.add_argument('--lr', type=float, default=1e-4)
    parser.add_argument('--save_interval', type=int, default=100)
    parser.add_argument('--model', type=str, default='Cnn')
    parser.add_argument('--obj_dir', metavar='PATH', default='../results/classification accuracy')
    args = parser.parse_args()
    args.save_dir = args.obj_dir

    data_dir = Path(
        './results/')  # D:/Data/DLSpace/results/path2pose ; /data/cjj/DLSpace/results/path2pose/
    mode_list = [
        'AttnCn10000', 'AttnCn2000', 'AttnCn30000', 'AttnCn40000', 'AttnCn50000', 'AttnCn60000',
        'Gcn10000', 'Gcn20000', 'Gcn30000', 'Gcn40000',
        'Rnn10000', 'Rnn20000', 'Rnn30000', 'Rnn40000',
    ]  # todo
    npz_list = [  # todo
        'AttnCnNet/val/epoch20000/test_results.npz',
        'Gcn/val/epoch20000/test_results.npz',
        'RNN/val/epoch20000/test_results.npz',
    ]

    assert len(mode_list) == len(npz_list)
    # assert [mode_list[i] in npz_list[i] for i in range(len(mode_list))]

    model_pool = {'Rnn': Classifier_Rnn,
                  'Cnn': Classifier_Cnn_40}
    print('model:' + args.model)

    args.save_dir = Path(args.save_dir)
    if not args.save_dir.exists():
        args.save_dir.mkdir(parents=True)

    args.log_train = args.save_dir / 'log_train.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append(['epoch'] + mode_list)
    wb.save(args.log_train)

    args.log_test = args.save_dir / 'log_test.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append(['epoch'] + mode_list)
    wb.save(args.log_test)

    np.random.seed(1234)
    torch.manual_seed(1234)
    torch.cuda.manual_seed(1234)

    args.device = torch.device('cuda:3')
    args.w = torch.tensor([1, 1]).float().to(args.device)  # todo

    for i in range(len(mode_list)):
        train_pose, train_label, train_detail_label, test_pose, test_label, test_detail_label, _ = \
            build_data(data_dir, [npz_list[i]], enhance=True, train_rate=0.8)  # todo 是否进行增强， 比率

        train_loader = prepare_dataloader(train_pose, train_label, train_detail_label, args.batch_size)
        test_loader = prepare_dataloader(test_pose, test_label, test_detail_label, args.batch_size)

        model = model_pool[args.model]()
        model = model.to(args.device)
        train(train_loader, test_loader, model, i, args)
        del model
